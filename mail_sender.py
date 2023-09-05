import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
import openpyxl


# Email configuration
sender_email = "your email here"
sender_password = "your app password" #Under "Signing in to Google," select 2-Step Verification. At the bottom of the page, select App passwords.

# Load data from Excel file
excel_file_path = "Email_List.xlsx"  # Replace with the path to your Excel file
data = pd.read_excel(excel_file_path, sheet_name='Sheet1')  # Assuming data is in Sheet1

# Check if the required columns exist
if 'Email' not in data.columns or 'Name' not in data.columns:
    print("Error: Required columns (Email and/or Name) not found in the Excel file.")
    exit()

# Convert the data into a list of tuples (email, name, status)
recipient_list = [(row['Email'], row['Name'], row['Status']) for index, row in data.iterrows()]



# Email content
message_template = """
<your content here. this is HTML format. if you want use plain text, change this code msg.attach(MIMEText(message, 'html')) as 'msg.attach(MIMEText(message, 'plain'))'>

"""

# Connect to the Gmail SMTP server
smtp_server = "smtp.gmail.com"
smtp_port = 587

# Constant attachment filename
filename = "you can add your attachment file here"  # Change to your constant filename

try:
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)
    
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Sheet1']  # Change to the sheet containing data
    
    for recipient_email, recipient_name, status in recipient_list:
        # Check if the status is "completed"; if yes, skip sending the email
        if status == "completed":
            print(f"Email to {recipient_email} already marked as completed. Skipping.")
            continue
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = "Your subject here"  # Replace with your desired subject
        
        message = message_template.format(name=recipient_name, week_number="X")
        msg.attach(MIMEText(message, 'html'))
        
        # Attach the constant file name
        attachment = open(filename, "rb")
        
        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(part)
        
        server.sendmail(sender_email, recipient_email, msg.as_string())
        print(f"Email with attachment sent to {recipient_email}")
        
        # Update the status to "completed"
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            email, name = [cell.value for cell in row]
            if email == recipient_email and name == recipient_name:
                status_cell = sheet.cell(row=row[0].row, column=data.columns.get_loc('Status') + 1)
                status_cell.value = "completed"  # Set the status in the 'Status' column
                break
    
    server.quit()
    
    # Save the updated Excel file
    wb.save(excel_file_path)

except Exception as e:
    print(f"Error: {e}")